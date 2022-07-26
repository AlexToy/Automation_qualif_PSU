import sys
import pyvisa
import openpyxl
from openpyxl.chart import LineChart, Reference
import time
import os

debug = 0

'''
- Waring: The disjunction detection is triggered when a abnormal Vmin value is read.
-         If the voltage is goes out of the scope of the screen, the software will detect a disjunction while
-         is it possible that the PSU is not in disjunction mode.
'''


class Psu:
    def __init__(self):
        self.name = None
        self.ref = None
        self.part_num = None
        self.cie = None

    def set_infos(self):
        print("PSU name :")
        self.name = input()
        print("PSU ref :")
        self.ref = input()
        print("PSU part number :")
        self.part_num = input()
        print("PSU cie :")
        self.cie = input()

    def print_infos(self):
        print(self.name)
        print(self.ref)
        print(self.part_num)
        print(self.cie)


class Measures:
    def __init__(self):
        self.power = None
        self.current_avrg = None
        self.current_max = None
        self.voltage_avrg = None
        self.voltage_min = None
        self.screenshot_name = None
        self.disjunction = None

    def print_measure(self):
        print("Current Avrg : " + str(self.current_avrg))
        print("Current max : " + str(self.current_max))
        print("Voltage Avrg : " + str(self.voltage_avrg))
        print("Voltage min : " + str(self.voltage_min))


class GroupByFrequency:
    def __init__(self, frequency, h_scale, current_max_test):
        self.frequency = frequency
        self.list_measures = []
        self.h_scale = h_scale
        self.screenshot_index = 0
        self.current_max_test = current_max_test

    def set_frequency(self, oscilloscope_session, active_load_session):
        # Set the active load frequency
        active_load_session.write('CURR:SINE:FREQ ' + self.frequency)

        # Modify the horizontal scale to be compliant with the frequency
        print("Setup scope for " + self.frequency)
        oscilloscope_session.write(self.h_scale)
        print("Setup done")

    def find_disjunction(self, duration_test, session):
        for _ in range(duration_test):
            voltage = session.query_ascii_values(":MEASure:VMIN? CHANnel2")[0]
            if voltage > 100:
                print("Disjunction")
                return True
            time.sleep(1)
        return False

    def add_measures(self, oscilloscope_session, active_load_session, folder_name):
        print("--- New test frequency : " + self.frequency + " Hz ---")
        current_avrg_at = 1
        current_amp_at = 2
        disjunction = False

        while not disjunction:  # The current max for the test must be set at the beginning
            print("--> New measure : Current Average = " + str(current_avrg_at) + " A ; Current Amplitude = ",
                  str(current_amp_at) + " A")
            # 1 : Set the active load with its parameters
            active_load_session.write('INIT:SINE')
            active_load_session.write('CURR ' + str(current_avrg_at))
            active_load_session.write('CURR:SINE:AMPL ' + str(current_amp_at))
            active_load_session.write('OUTPut ON')
            time.sleep(3)

            # 2 : Take the measures
            new_measures = Measures()
            new_measures.current_avrg = oscilloscope_session.query_ascii_values(":MEASure:VAVerage? DISPlay,CHANnel4")[0]
            new_measures.current_max = oscilloscope_session.query_ascii_values(":MEASure:VMAX? CHANnel4")[0]
            new_measures.voltage_avrg = oscilloscope_session.query_ascii_values(":MEASure:VAVerage? DISPlay,CHANnel2")[0]
            new_measures.voltage_min = oscilloscope_session.query_ascii_values(":MEASure:VMIN? CHANnel2")[0]
            new_measures.screenshot_name = str(self.frequency) + str(self.screenshot_index)
            new_measures.power = round((float(new_measures.voltage_avrg) * float(new_measures.current_avrg)), 0)
            self.list_measures.append(new_measures)
            # Print the measurements
            new_measures.print_measure()

            # 3 : Take a screenshot
            take_screenshot(oscilloscope_session, str(self.frequency) + "_" + str(self.screenshot_index), folder_name)

            # 4 : There is a disjunction ?
            if current_amp_at < self.current_max_test:
                disjunction = self.find_disjunction(5, oscilloscope_session)
                new_measures.disjunction = disjunction
            else:
                # If the current max is at max current test stop measurements at this frequency
                new_measures.disjunction = False
                disjunction = True

            time.sleep(1)
            active_load_session.write('ABOR:SINE')
            active_load_session.write('OUTPut OFF')
            time.sleep(1)

            # Update the current values
            current_avrg_at = current_avrg_at + 0.5
            current_amp_at = current_amp_at + 1
            self.screenshot_index = self.screenshot_index + 1


def take_screenshot(oscilloscope_session, file_name, folder_name):
    oscilloscope_session.timeout = 4000
    os.makedirs(folder_name, exist_ok=True)
    screen_bytes = do_query_ieee_block(oscilloscope_session, ":DISPlay:DATA? PNG")
    # Save display data values to file.
    file = str(folder_name) + "/" + str(file_name) + ".png"
    f = open(file, "wb")
    f.write(screen_bytes)
    f.close()
    print("Screen image written to " + file)


def check_instrument_errors(oscilloscope_session, command, exit_on_error=True):
    while True:
        error_string = oscilloscope_session.query(":SYSTem:ERRor? STRing")
        if error_string: # If there is an error string value.
            if error_string.find("0,", 0, 2) == -1: # Not "No error".
                print("ERROR: %s, command: '%s'" % (error_string, command))
                if exit_on_error:
                    print("Exited because of error.")
                    sys.exit(1)
            else: # "No error"
                break
        else: # :SYSTem:ERRor? STRing should always return string.
            print("ERROR: :SYSTem:ERRor? STRing returned nothing, command: '%s'" % command)
            print("Exited because of error.")
            sys.exit(1)


def do_query_ieee_block(oscilloscope_session, query):
    if debug:
        print("Qyb = '%s'" % query)
    result = oscilloscope_session.query_binary_values("%s" % query, datatype='s', container = bytes)
    check_instrument_errors(oscilloscope_session, query, exit_on_error=False)
    return result


def create_graphics(working_sheet, working_frequency, first_chart_row, last_chart_row, graphics_init_pos):
    # Create the graphic
    power = Reference(working_sheet, min_col=1, min_row=first_chart_row, max_row=last_chart_row)
    voltage_min = Reference(working_sheet, min_col=5, min_row=first_chart_row - 1, max_row=last_chart_row)
    current_avrg = Reference(working_sheet, min_col=2, min_row=first_chart_row - 1, max_row=last_chart_row)
    current_max = Reference(working_sheet, min_col=3, min_row=first_chart_row - 1, max_row=last_chart_row)

    # First line chart : Voltage min
    voltage_min_line = LineChart()
    voltage_min_line.title = "Sinus load " + str(working_frequency) + "Hz"
    voltage_min_line.x_axis.title = "Puissance"
    voltage_min_line.y_axis.title = "Tension"
    voltage_min_line.y_axis.majorGridlines = None
    voltage_min_line.add_data(voltage_min, titles_from_data=True)
    voltage_min_line.set_categories(power)

    # Second line chart : Current Avrg
    current_avrg_line = LineChart()
    current_avrg_line.y_axis.title = "Courant"
    current_avrg_line.y_axis.axId = 200
    current_avrg_line.add_data(current_avrg, titles_from_data=True)

    # Third line chart : Current Avrg
    current_max_line = LineChart()
    current_max_line.y_axis.title = "Current Max"
    current_max_line.add_data(current_max, titles_from_data=True)

    current_avrg_line.y_axis.crosses = "max"
    voltage_min_line += current_avrg_line
    voltage_min_line += current_max_line

    working_sheet.add_chart(voltage_min_line, "k" + str(graphics_init_pos))


def create_final_graphic(working_sheet, first_chart_row, last_chart_row):
    frequency = Reference(working_sheet, min_col=1, min_row=first_chart_row + 1, max_row=last_chart_row)
    current_avrg = Reference(working_sheet, min_col=2, min_row=first_chart_row, max_row=last_chart_row)
    current_max = Reference(working_sheet, min_col=3, min_row=first_chart_row, max_row=last_chart_row)

    # First line chart : Current Avrg
    current_avrg_line = LineChart()
    current_avrg_line.title = "Disjunction current according frequency from sinus load"
    current_avrg_line.x_axis.title = "Frequency"
    current_avrg_line.y_axis.title = "Current"
    current_avrg_line.add_data(current_avrg, titles_from_data=True)
    current_avrg_line.set_categories(frequency)

    # Second line chart : Current Max
    current_max_line = LineChart()
    current_max_line.add_data(current_max, titles_from_data=True)

    current_avrg_line += current_max_line

    working_sheet.add_chart(current_avrg_line, "k" + str(first_chart_row))


def excel(psu, list_grp_frequency, folder_name):
    workbook = openpyxl.Workbook()
    name_file = str(folder_name) + "/measurements.xlsx"

    # 1 : Write PSU information
    sheet = workbook.active
    sheet['A1'] = 'PSU'
    sheet['B1'] = psu.name
    sheet['A2'] = 'Ref'
    sheet['B2'] = psu.ref
    sheet['A3'] = 'Part Num.'
    sheet['B3'] = psu.part_num
    sheet['A4'] = 'CIE.'
    sheet['B4'] = psu.cie

    # 2 : Write all the measurement chart
    row = 7
    for grp_frequency in list_grp_frequency:
        # Create the title line of the chart
        sheet.cell(row, 1, "Sinus " + str(grp_frequency.frequency))
        row = row + 1
        sheet.cell(row, 1, "Puissance")
        sheet.cell(row, 2, "Courant Avrg")
        sheet.cell(row, 3, "Courant Max")
        sheet.cell(row, 4, "Tension Avrg")
        sheet.cell(row, 5, "Tension Min")
        sheet.cell(row, 6, "Screenshot name")
        sheet.cell(row, 7, "Disjunction")

        # Set the original position of the graphics
        graphics_init_pos = row

        row = row + 1
        first_chart_row = row
        for measures in grp_frequency.list_measures:
            sheet.cell(row, 1, measures.power)
            sheet.cell(row, 2, measures.current_avrg)
            sheet.cell(row, 3, measures.current_max)
            sheet.cell(row, 4, measures.voltage_avrg)
            sheet.cell(row, 5, measures.voltage_min)
            sheet.cell(row, 6, measures.screenshot_name)
            sheet.cell(row, 7, measures.disjunction)
            row = row + 1

    # 3 : Create and print all graphics
        # Remove the last row if it is a disjunction row
        if grp_frequency.list_measures[-1].disjunction:
            last_chart_row = row - 2
        else:
            last_chart_row = row - 1
        create_graphics(sheet, grp_frequency.frequency, first_chart_row, last_chart_row, graphics_init_pos)
        row = row + 2   # Separation between two charts

    # 4 : Create the final chart and display its graph
    sheet.cell(row, 1, "Frequence")
    sheet.cell(row, 2, "Courant Avrg")
    sheet.cell(row, 3, "Courant Max")
    first_final_chart_row = row
    row = row + 1

    for grp_frequency in list_grp_frequency:
        print(grp_frequency.frequency)
        sheet.cell(row, 1, grp_frequency.frequency)
        # If the last measure there was a disjunction, write second to last current (not the disjunction current).
        # Else write the last current
        if grp_frequency.list_measures[-1]:
            print("True")
            sheet.cell(row, 2, grp_frequency.list_measures[-2].current_avrg)
            sheet.cell(row, 3, grp_frequency.list_measures[-2].current_max)
        else:
            print("False")
            sheet.cell(row, 2, grp_frequency.list_measures[-1].current_avrg)
            sheet.cell(row, 3, grp_frequency.list_measures[-1].current_max)
        row = row + 1

    last_final_shart_row = row - 1
    create_final_graphic(sheet, first_final_chart_row, last_final_shart_row)

    workbook.save(name_file)
    workbook.close()


def create_connection_oscilloscope():
    print("Oscilloscope IP Address : ")
    # ip = input()
    ip = '10.80.64.133'
    visa_address = 'TCPIP::' + ip + '::inst0::INSTR'

    try:
        # Create a connection (session) to the instrument
        resource_manager = pyvisa.ResourceManager()
        session = resource_manager.open_resource(visa_address)
        print("Connection success !")
        return session
    except pyvisa.Error as ex:
        print('Couldn\'t connect to \'%s\', exiting now...' % visa_address)
        sys.exit()


def create_connection_load():
    print("Active load IP Address : ")
    # ip = input()
    ip = '10.80.64.137'
    visa_address = 'TCPIP::' + ip + '::inst0::INSTR'

    try:
        # Create a connection (session) to the instrument
        resource_manager = pyvisa.ResourceManager()
        session = resource_manager.open_resource(visa_address)
        print("Connection success !")
        return session
    except pyvisa.Error as ex:
        print('Couldn\'t connect to \'%s\', exiting now...' % visa_address)
        sys.exit()


# List of different frequencies that must be tested and the correspondence of the horizontal scale
# list_frequency = [["1000", ":TIMebase:SCALe 2E-3"], ["2000", ":TIMebase:SCALe 1E-3"]]


list_frequency = [["200", ':TIMebase:SCALe 5E-3'], ["500", ":TIMebase:SCALe 5E-3"],
                  ["1000", ":TIMebase:SCALe 2E-3"], ["2000", ":TIMebase:SCALe 1E-3"],
                  ["5000", ":TIMebase:SCALe 5E-4"], ["10000", ":TIMebase:SCALe 2E-4"]]


def main():
    # 1 : User information
    print("--- PROB SETUP ---")
    print("Voltage prob : must be plug on channel 2")
    print("   - PSU 12V : Vertical scale --> 1 V")
    print("               Offset         --> 15.7 V")
    print("   - PSU 16V : Vertical scale --> --- mV")
    print("               Offset         --> --- V")
    print(" ")
    print("Current prob : must be plug on channel 4")
    print("Please, calibrate the prob and verify and check the direction !")
    print("   - PSU 12V : Vertical scale --> 2 A")
    print("               Offset         --> 6 A")
    print("   - PSU 16V : Vertical scale --> 2 A")
    print("               Offset         --> 6 A")
    print(" ")
    print("Horizontal scale is automatically updated !")
    print("When your are ready, click on 'Y' ")
    input()

    # 2 : Create connection with the scope and the active load
    oscilloscope_session = create_connection_oscilloscope()
    active_load_session = create_connection_load()

    # 3 : Get information about PSU
    print("Enter PSU information")

    psu = Psu()
    psu.set_infos()

    # 4 : Get current maximum to test
    print("Enter the maximum current to test")
    current_max_test = float(input())

    # 5 : get folder name for screenshot
    print("Folder name ?")
    folder_name = input()

    # 6 : Create a new object GroupByFrequency for each frequency that is in the list
    list_group_by_frequency = []
    for item in list_frequency:
        # GroupByFrequency( frequency, command for oscilloscope horizontal scale )
        list_group_by_frequency.append(GroupByFrequency(item[0], item[1], current_max_test))

    # 7 : For each frequency, takes measurements
    for group_by_frequency in list_group_by_frequency:
        group_by_frequency.set_frequency(oscilloscope_session, active_load_session)
        group_by_frequency.add_measures(oscilloscope_session, active_load_session, folder_name)

    # 8 : Write measurements and crete all graphics in a excel file
    excel(psu, list_group_by_frequency, folder_name)


main()
